# excelizing.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
import os
from datetime import datetime

def generate_timetable(currentschedule, lectures, professors, classrooms, file_path):
    def visualization(currentschedule, lectures, professors, classrooms):
        # 템플릿 로딩
        file_path_template = "C:/Users/PC/Desktop/pymongo/algorithm/template.xlsx"
        wb = load_workbook(filename=file_path_template)
        row_height = 54.75  # 추가될 행의 높이

        ## 학년명 레이블 추가
        headers = ["교양", "1학년", "2학년", "3학년", "4학년", "대학원"]
        start_column = column_index_from_string('E')

        # 색상환
        color_list = ["BFBFBF","E6B8B7","FCD5B5","92CDDD","D7E4BC","FFFF66"]
        batch_color_list = ["D9D9D9","F3DCDB","FCD5B5","C6DAF1","D7E4BC","FFFF66"]
        batch_color_list_bold = ["D9D9D9","D99694","E46C0A","568ED4","77933C","FFFF66"]
        color_dict = {}

        # 강의별 색상 지정
        for lecture in lectures:
            lectureCode = lecture.lectureCode
            year = lecture.year
            baseColor = batch_color_list[year]

            if lectureCode not in color_dict:
                if lecture.MR:
                    baseColor = batch_color_list_bold[year]

                # color_dict에 새로운 색상 추가
                color_dict[lectureCode] = baseColor

        for sheet in wb.worksheets:
            # 기존 병합 해제 (병합 범위를 리스트로 복사하여 해제)
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                sheet.unmerge_cells(str(merged_range))
            
            # 맨 위에 새로운 행 추가
            sheet.insert_rows(1)

            # 학년명 레이블과 색상 추가
            for i, (header, color) in enumerate(zip(headers, color_list)):
                col = start_column + i * 4  # 학년마다 4열 간격
                
                # 색상 추가 (각 레이블의 왼쪽 셀에 색상 채우기)
                color_cell = sheet.cell(row=1, column=col - 2)
                color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # 학년명 레이블 추가
                label_cell = sheet.cell(row=1, column=col)
                label_cell.value = header
                label_cell.alignment = Alignment(horizontal="center", vertical="center")

        ## 교원 시간표
        sheet = wb.worksheets[0]
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        start_row = 4
        start_column = column_index_from_string('B')

        for professor in professors:
            sheet[f'A{start_row}'] = professor.name
            sheet[f'A{start_row}'].alignment = Alignment(horizontal="center", vertical="center")

            profCode = professor.profCode

            for lecture in currentschedule:
                if lecture.profCode == profCode:
                    day = lecture.batched[0]
                    period = lecture.batched[1]
                    duration = lecture.duration

                    for d in range(duration):
                        lecture_column = get_column_letter(start_column + day * 13 + period + d)
                        lecture_cell = f"{lecture_column}{start_row}"
                        sheet[lecture_cell] = f"{lecture.name}-{lecture.division}({professor.name})"
                        sheet[lecture_cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        if lecture.lectureCode in color_dict:
                            color = color_dict[lecture.lectureCode]
                            sheet[lecture_cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # 행 높이 설정
            sheet.row_dimensions[start_row].height = row_height

            # 헤더 종합
            for i in range(5):
                # 병합할 시작 및 끝 열 계산
                col_start = 2 + i * 13
                col_end = col_start + 13 - 1

                # 열을 문자로 변환 후 병합
                merge_range = f"{get_column_letter(col_start)}2:{get_column_letter(col_end)}2"
                sheet.merge_cells(merge_range)

                # 병합된 셀의 중앙 정렬 설정
                cell = sheet.cell(row=2, column=col_start)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet[f'A{2}'].alignment = Alignment(horizontal="right", vertical="center")
            sheet[f'A{3}'].alignment = Alignment(horizontal="center", vertical="center")

            # 경계선 설정
            sheet.cell(row=start_row, column=1).border = Border( # A열
                    bottom=Side(style="thin"),
                    right=Side(style="thin")
                )
            for col in range(2, sheet.max_column): # 밑줄
                sheet.cell(row=start_row, column=col).border = Border(
                    bottom=Side(style="thin")
                )
            for col in range(1, sheet.max_column + 1, 13): # 날짜줄
                sheet.cell(row=start_row, column=col).border = Border(
                    bottom=Side(style="thin"),
                    right=Side(style="thin")
                )
            
            # 폰트 설정
            for col in range(2, sheet.max_column + 1):
                sheet.cell(row=start_row, column=col).font = Font(size=8)

            # 다음 행 이동
            start_row += 1

        # 교원 시간표 병합
        max_row = sheet.max_row
        max_col = sheet.max_column
        for row in range(3, max_row + 1):
            start_col = 2
            while start_col < max_col:
                cell = sheet.cell(row=row, column=start_col)
                if cell.value:
                    end_col = start_col
                    while end_col + 1 <= max_col and sheet.cell(row=row, column=end_col + 1).value == cell.value:
                        end_col += 1

                    if end_col > start_col:
                        sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

                    start_col = end_col + 1
                else:
                    start_col += 1

        ## 학년 시간표
        sheet = wb.worksheets[1]

        # 시트 전처리
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet[f'A{1}'].alignment = Alignment(horizontal="right", vertical="center")
        sheet[f'A{2}'].alignment = Alignment(horizontal="center", vertical="center")
        start_row = 4  # 학년별 시작 행
        start_column = column_index_from_string('B')

        for i in range(0, 6):
            # 학년 또는 분류 제목 추가
            if i == 0:
                title = "교양"
            elif i == 5:
                title = "대학원"
            else:
                title = f"{i}학년"
            
            # 학년 이름을 첫 번째 열에 배치
            sheet[f'A{start_row}'] = title
            sheet[f'A{start_row}'].alignment = Alignment(horizontal="center", vertical="center")

            current_row = start_row  # 현재 강의 배치를 위한 시작 행
            max_row_grade = start_row  # 학년별 최대 행을 추적

            for lecture in currentschedule:
                if lecture.year == i:
                    # 교원명 탐색
                    profName = None
                    for professor in professors:
                        if professor.profCode == lecture.profCode:
                            profName = professor.name
                            break

                    day = lecture.batched[0]
                    period = lecture.batched[1]
                    duration = lecture.duration

                    for d in range(duration):
                        lecture_column = get_column_letter(start_column + day * 13 + period + d)
                        lecture_cell = f"{lecture_column}{current_row}"

                        # 셀에 값이 이미 있는 경우, 다음 행으로 이동
                        while sheet[lecture_cell].value:
                            # 경계선 설정
                            sheet.cell(row=current_row, column=1).border = Border( # A열
                                    bottom=Side(style="thin"),
                                    right=Side(style="thin")
                                )
                            for col in range(2, sheet.max_column): # 밑줄
                                sheet.cell(row=current_row, column=col).border = Border(
                                    bottom=Side(style="thin")
                                )
                            for col in range(1, sheet.max_column + 1, 13): # 날짜줄
                                sheet.cell(row=current_row, column=col).border = Border(
                                    bottom=Side(style="thin"),
                                    right=Side(style="thin")
                                )
                            current_row += 1
                            sheet.row_dimensions[current_row].height = row_height
                            lecture_cell = f"{lecture_column}{current_row}"

                            if lecture.lectureCode in color_dict:
                                color = color_dict[lecture.lectureCode]
                                sheet[lecture_cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                        # 셀이 비어 있는 경우 값을 할당
                        sheet[lecture_cell] = f"{lecture.name}-{lecture.division}({profName})"
                        sheet[lecture_cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        if lecture.lectureCode in color_dict:
                            color = color_dict[lecture.lectureCode]
                            sheet[lecture_cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                    max_row_grade = max(max_row_grade, current_row)
                    current_row = start_row

            # 학년별 A열 병합
            if max_row_grade > start_row:
                sheet.merge_cells(start_row=start_row, start_column=1, end_row=max_row_grade, end_column=1)

            # 행 높이 설정
            sheet.row_dimensions[start_row].height = row_height

            # 헤더 종합
            for i in range(5):
                # 병합할 시작 및 끝 열 계산
                col_start = 2 + i * 13
                col_end = col_start + 13 - 1

                # 열을 문자로 변환 후 병합
                merge_range = f"{get_column_letter(col_start)}2:{get_column_letter(col_end)}2"
                sheet.merge_cells(merge_range)

                # 병합된 셀의 중앙 정렬 설정
                cell = sheet.cell(row=2, column=col_start)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet[f'A{2}'].alignment = Alignment(horizontal="right", vertical="center")
            sheet[f'A{3}'].alignment = Alignment(horizontal="center", vertical="center")
            
            # 경계선 설정
            for col in range(2, sheet.max_column): # 밑줄
                sheet.cell(row=max_row_grade, column=col).border = Border(
                    bottom=Side(style="thin")
                )
            for col in range(1, sheet.max_column + 1, 13): # 날짜줄
                sheet.cell(row=max_row_grade, column=col).border = Border(
                    bottom=Side(style="thin"),
                    right=Side(style="thin")
                )

            # 폰트 설정
            for row in range(4, max_row_grade + 1):
                for col in range(2, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).font = Font(size=8)

            # 다음 행 이동
            start_row = max_row_grade + 1

        # 모든 셀을 검토하며 같은 내용이 연속적으로 있는 경우 병합
        max_row = sheet.max_row
        max_col = sheet.max_column
        for row in range(3, max_row + 1):  # 시작 행부터 마지막 행까지
            start_col = 2  # 시작 열 (B열)
            while start_col < max_col:
                cell = sheet.cell(row=row, column=start_col)
                if cell.value:
                    # 같은 내용을 가진 셀의 범위를 탐색
                    end_col = start_col
                    while end_col + 1 <= max_col and sheet.cell(row=row, column=end_col + 1).value == cell.value:
                        end_col += 1

                    # 동일한 값을 가진 셀들을 병합
                    if end_col > start_col:  # 범위가 2칸 이상인 경우에만 병합
                        sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

                    # 다음 병합 범위 탐색을 위해 start_col을 갱신
                    start_col = end_col + 1
                else:
                    # 현재 셀이 비어있으면 다음 셀로 이동
                    start_col += 1

        ## 강의실 시간표
        sheet = wb.worksheets[2]
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet[f'A{1}'].alignment = Alignment(horizontal="right", vertical="center")
        sheet[f'A{2}'].alignment = Alignment(horizontal="center", vertical="center")

        start_row = 4
        start_column = column_index_from_string('C')

        for classroom in classrooms:
            building = classroom.building
            classroomNo = classroom.classroomNo

            sheet[f'A{start_row}'] = building
            sheet[f'A{start_row}'].alignment = Alignment(horizontal="center", vertical="center")
            sheet[f'B{start_row}'] = classroomNo
            sheet[f'B{start_row}'].alignment = Alignment(horizontal="center", vertical="center")

            for lecture in currentschedule:
                if lecture.batched[2] == building and lecture.batched[3] == classroomNo:
                    # 교원명 탐색
                    profName = None
                    for professor in professors:
                        if professor.profCode == lecture.profCode:
                            profName = professor.name
                            break

                    day = lecture.batched[0]
                    period = lecture.batched[1]
                    duration = lecture.duration

                    for d in range(duration):
                        lecture_column = get_column_letter(start_column + day * 13 + period + d)
                        lecture_cell = f"{lecture_column}{start_row}"
                        sheet[lecture_cell] = f"{lecture.name}-{lecture.division}({profName})"
                        sheet[lecture_cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        if lecture.lectureCode in color_dict:
                            color = color_dict[lecture.lectureCode]
                            sheet[lecture_cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # 행 높이 설정
            sheet.row_dimensions[start_row].height = row_height

            # 헤더 종합
            for i in range(5):
                # 병합할 시작 및 끝 열 계산
                col_start = 3 + i * 13
                col_end = col_start + 13 - 1

                # 열을 문자로 변환 후 병합
                merge_range = f"{get_column_letter(col_start)}2:{get_column_letter(col_end)}2"
                sheet.merge_cells(merge_range)

                # 병합된 셀의 중앙 정렬 설정
                cell = sheet.cell(row=2, column=col_start)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.merge_cells("A2:B2")
            sheet[f'A{2}'].alignment = Alignment(horizontal="right", vertical="center")
            sheet[f'A{3}'].alignment = Alignment(horizontal="center", vertical="center")
            sheet[f'B{3}'].alignment = Alignment(horizontal="center", vertical="center")

            # 경계선 설정
            sheet.cell(row=start_row, column=1).border = Border( # A열
                    bottom=Side(style="thin"),
                    right=Side(style="thin")
                )
            sheet.cell(row=start_row, column=2).border = Border( # B열
                    bottom=Side(style="thin"),
                    right=Side(style="thin")
                )
            for col in range(3, sheet.max_column): # 밑줄
                sheet.cell(row=start_row, column=col).border = Border(
                    bottom=Side(style="thin")
                )
            for col in range(2, sheet.max_column + 1, 13): # 날짜줄
                sheet.cell(row=start_row, column=col).border = Border(
                    bottom=Side(style="thin"),
                    right=Side(style="thin")
                )
            
            # 폰트 설정
            for col in range(3, sheet.max_column + 1):
                sheet.cell(row=start_row, column=col).font = Font(size=8)

            # 다음 행 이동
            start_row += 1

        # 강의실 시간표 병합
        max_row = sheet.max_row
        max_col = sheet.max_column
        for row in range(3, max_row + 1):
            start_col = 3
            while start_col < max_col:
                cell = sheet.cell(row=row, column=start_col)
                if cell.value:
                    end_col = start_col
                    while end_col + 1 <= max_col and sheet.cell(row=row, column=end_col + 1).value == cell.value:
                        end_col += 1

                    if end_col > start_col:
                        sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

                    start_col = end_col + 1
                else:
                    start_col += 1
        
        return wb

    # 시간표 정렬
    currentschedule.sort(key=lambda x: (x.lectureCode, x.division, x.TP))
    professors = sorted(professors, key=lambda professor: not professor.isprof)

    # 시간표 생성 및 저장
    wb = visualization(currentschedule, lectures, professors, classrooms)

    # 파일 저장
    wb.save(file_path)
    print("시간표 생성 및 엑셀 파일 저장이 완료되었습니다.")

    return file_path
